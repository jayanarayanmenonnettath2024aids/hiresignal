"""
ExportService: Generate CSV and Excel exports of resume results.
"""

import io
from typing import List, BinaryIO
from uuid import UUID
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models import ResumeResult, ScreeningJob


class ExportService:
    """Generate CV/Excel exports"""

    CSV_HEADERS = [
        'Rank',
        'Filename',
        'Final Score',
        'Semantic Score',
        'TF-IDF Score',
        'Skill Match Score',
        'Experience Score',
        'Matched Skills',
        'Missing Skills',
        'YoE Detected',
        'Language',
        'Flags',
        'Feedback Action',
    ]
    
    @staticmethod
    async def export_csv(
        db: AsyncSession,
        job_id: UUID,
        results: List[ResumeResult]
    ) -> bytes:
        """
        Export results to CSV.
        Scores as 0-100 (multiply by 100).
        """
        job = await ExportService._get_job(db, job_id)
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=ExportService.CSV_HEADERS)
        
        writer.writeheader()
        
        ordered_results = sorted(
            results,
            key=lambda r: (
                r.rank is None,
                r.rank if r.rank is not None else 10**9,
                -(r.final_score or 0.0),
            ),
        )

        for result in ordered_results:
            writer.writerow({
                'Rank': result.rank or '',
                'Filename': result.filename,
                'Final Score': f"{result.final_score * 100:.1f}" if result.final_score is not None else '',
                'Semantic Score': f"{result.score_semantic * 100:.1f}" if result.score_semantic is not None else '',
                'TF-IDF Score': f"{result.score_tfidf * 100:.1f}" if result.score_tfidf is not None else '',
                'Skill Match Score': f"{result.score_skills * 100:.1f}" if result.score_skills is not None else '',
                'Experience Score': f"{result.score_experience * 100:.1f}" if result.score_experience is not None else '',
                'Matched Skills': ', '.join(result.matched_skills),
                'Missing Skills': ', '.join(result.missing_skills),
                'YoE Detected': f"{result.years_experience_detected:.1f}" if result.years_experience_detected is not None else '',
                'Language': result.language_detected,
                'Flags': ', '.join(result.flags.keys()),
                'Feedback Action': ''
            })
        
        return output.getvalue().encode('utf-8')
    
    @staticmethod
    async def export_excel(
        db: AsyncSession,
        job_id: UUID,
        results: List[ResumeResult]
    ) -> bytes:
        """
        Export results to Excel with multiple sheets.
        Sheet 1: Shortlist
        Sheet 2: Score Breakdown
        Sheet 3: JD Info
        """
        job = await ExportService._get_job(db, job_id)
        
        wb = Workbook()
        
        # Sheet 1: Shortlist
        ws = wb.active
        ws.title = "Shortlist"
        ExportService._write_shortlist_sheet(ws, results)
        
        # Sheet 2: Score Breakdown
        ws2 = wb.create_sheet("Score Breakdown")
        ExportService._write_breakdown_sheet(ws2, results)
        
        # Sheet 3: JD Info
        ws3 = wb.create_sheet("JD Info")
        ExportService._write_jd_sheet(ws3, job)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def _write_shortlist_sheet(ws, results: List[ResumeResult]):
        """Write shortlist to sheet with formatting"""
        headers = [
            'Rank', 'Filename', 'Final Score (0-100)', 
            'Semantic', 'TF-IDF', 'Skills', 'Experience',
            'Matched Skills', 'Missing Skills', 'YoE Detected',
            'Language', 'Flags'
        ]
        
        ws.append(headers)
        
        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        
        ordered_results = sorted(
            results,
            key=lambda r: (
                r.rank is None,
                r.rank if r.rank is not None else 10**9,
                -(r.final_score or 0.0),
            ),
        )

        for result in ordered_results:
            score_100 = result.final_score * 100 if result.final_score else 0
            row = [
                result.rank or '',
                result.filename,
                f"{score_100:.1f}",
                f"{result.score_semantic * 100:.1f}" if result.score_semantic else '',
                f"{result.score_tfidf * 100:.1f}" if result.score_tfidf else '',
                f"{result.score_skills * 100:.1f}" if result.score_skills else '',
                f"{result.score_experience * 100:.1f}" if result.score_experience else '',
                ', '.join(result.matched_skills),
                ', '.join(result.missing_skills),
                f"{result.years_experience_detected:.1f}" if result.years_experience_detected else '',
                result.language_detected,
                ', '.join(result.flags.keys())
            ]
            ws.append(row)
            
            # Color code final score
            score_cell = ws[f'C{ws.max_row}']
            if result.final_score:
                if result.final_score >= 0.7:
                    score_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif result.final_score >= 0.4:
                    score_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
    
    @staticmethod
    def _write_breakdown_sheet(ws, results: List[ResumeResult]):
        """Write score breakdown"""
        ws.append(['Filename', 'Signal', 'Score'])
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for result in results:
            fs = result.filename
            if result.score_semantic:
                ws.append([fs, 'Semantic', f"{result.score_semantic:.3f}"])
            if result.score_tfidf:
                ws.append([fs, 'TF-IDF', f"{result.score_tfidf:.3f}"])
            if result.score_skills:
                ws.append([fs, 'Skills', f"{result.score_skills:.3f}"])
            if result.score_experience:
                ws.append([fs, 'Experience', f"{result.score_experience:.3f}"])
    
    @staticmethod
    def _write_jd_sheet(ws, job: ScreeningJob):
        """Write JD information"""
        ws.append(['Job Title', str(job.title)])
        ws.append(['Quality Score', f"{job.jd_quality_score:.3f}" if job.jd_quality_score else 'N/A'])
        ws.append(['Is Vague', 'Yes' if job.jd_is_vague else 'No'])
        ws.append([''])
        ws.append(['Extracted Skills'])
        
        for skill in job.jd_skills_extracted:
            ws.append(['', skill])
        
        ws.append([''])
        ws.append(['JD Text'])
        ws.append(['', job.jd_text])
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    @staticmethod
    async def _get_job(db: AsyncSession, job_id: UUID) -> ScreeningJob:
        """Get job details"""
        stmt = select(ScreeningJob).where(ScreeningJob.id == job_id)
        result = await db.execute(stmt)
        return result.scalar_one_or_none()


export_service = ExportService()
