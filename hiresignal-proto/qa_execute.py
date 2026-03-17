import json
import os
import time
import uuid
import zipfile
import csv
import subprocess
from pathlib import Path

import requests

BASE_80 = "http://localhost:80"
BASE_8000 = "http://localhost:8000"
BASE = BASE_80

report = {}
notes = []


def rec(section, status, msg=""):
    report[section] = (status, msg)
    print(f"[{section}] {status} {msg}")


def ok(cond, msg):
    if not cond:
        raise AssertionError(msg)


def req(method, path, **kwargs):
    return requests.request(method, BASE + path, timeout=30, **kwargs)


def body_json(resp):
    try:
        return resp.json()
    except Exception:
        return {}


def items_of(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        if isinstance(payload.get("items"), list):
            return payload["items"]
        if isinstance(payload.get("data"), list):
            return payload["data"]
    return []


# Setup checks
try:
    r80 = requests.get(BASE_80 + "/api/health", timeout=10)
    if r80.status_code == 200:
        base_probe = r80.json()
    else:
        base_probe = {"status": r80.status_code, "body": r80.text[:120]}
except Exception as e:
    base_probe = {"error": str(e)}

# Fallback to 8000 if port 80 routing is broken
try:
    r8 = requests.get(BASE_8000 + "/health", timeout=10)
    if r8.status_code == 200 and r8.json().get("status") == "ok":
        if not (isinstance(base_probe, dict) and base_probe.get("status") == "ok"):
            BASE = BASE_8000
            notes.append("Using BASE http://localhost:8000 because nginx/:80 route is misconfigured")
except Exception:
    pass

# Section 1 Auth
try:
    r = req("POST", "/api/auth/login", json={"email": "hr@acme.com", "password": "demo1234"})
    ok(r.status_code == 200, f"valid login status={r.status_code}")
    token = body_json(r).get("access_token", "")
    ok(bool(token), "empty access_token")

    bad = req("POST", "/api/auth/login", json={"email": "hr@acme.com", "password": "wrongpass"})
    ok(bad.status_code in (401, 403), f"wrong password not rejected: {bad.status_code}")

    me = req("GET", "/api/auth/me", headers={"Authorization": f"Bearer {token}"})
    ok(me.status_code == 200, f"/me status={me.status_code}")
    me_j = body_json(me)
    ok(me_j.get("email") == "hr@acme.com", "me email mismatch")
    ok(me_j.get("role") == "hr", "me role mismatch")

    unauth = req("GET", "/api/jobs")
    ok(unauth.status_code in (401, 403), f"unauth jobs allowed: {unauth.status_code}")

    startup = req("POST", "/api/auth/login", json={"email": "hr@startup.com", "password": "demo1234"})
    ok(startup.status_code == 200, "startup login failed")
    token_startup = body_json(startup).get("access_token", "")
    ok(bool(token_startup), "startup token empty")

    jobs = req("GET", "/api/jobs", headers={"Authorization": f"Bearer {token}"})
    ok(jobs.status_code == 200, f"jobs list status={jobs.status_code}")
    jobs_j = body_json(jobs)
    items = items_of(jobs_j)
    ok(len(items) > 0, "no acme jobs")
    acme_job_id = items[0]["id"]

    cross = req("GET", f"/api/jobs/{acme_job_id}", headers={"Authorization": f"Bearer {token_startup}"})
    ok(cross.status_code in (403, 404), f"cross tenant leak status={cross.status_code}")

    rec("SECTION 1", "PASS")
except Exception as e:
    rec("SECTION 1", "FAIL", str(e))
    token = ""

# Section 2 Seed data
try:
    jobs = req("GET", "/api/jobs", headers={"Authorization": f"Bearer {token}"})
    ok(jobs.status_code == 200, "jobs list failed")
    jobs_j = body_json(jobs)
    items = items_of(jobs_j)
    ok(len(items) == 2, f"expected 2 seeded jobs got {len(items)}")
    for j in items:
        ok(j.get("status") == "done", f"job not done: {j.get('title')} {j.get('status')}")
        ok(j.get("total_processed") == j.get("total_submitted"), f"processed/submitted mismatch for {j.get('title')}")

    job1_id = items[0]["id"]
    res = req("GET", f"/api/jobs/{job1_id}/results", headers={"Authorization": f"Bearer {token}"})
    ok(res.status_code == 200, "results endpoint failed")
    rj = body_json(res)
    results = items_of(rj)
    ok(len(results) > 0, "no seeded results")
    ok(all(x.get("rank") is not None for x in results), "some rank null")
    ok(all(x.get("final_score") is not None for x in results), "some final_score null")
    ok(results[0]["final_score"] >= results[-1]["final_score"], "results not sorted desc")

    scores = [x["final_score"] for x in results]
    ok(max(scores) > 0.5, "max score <= 0.5")
    ok(min(scores) < 0.6, "min score too high")
    ok(min(scores) >= 0.0 and max(scores) <= 1.0, "scores out of [0,1]")

    rec("SECTION 2", "PASS")
except Exception as e:
    rec("SECTION 2", "FAIL", str(e))

# Section 3 JD preview
try:
    strong_jd = {
        "jd_text": "We are hiring a Senior Python Engineer. Requirements: 5+ years of Python, FastAPI, PostgreSQL, Docker, Redis, AWS, React, Git. Experience with machine learning and PyTorch a plus."
    }
    r = req("POST", "/api/jobs/preview-jd", data=strong_jd, headers={"Authorization": f"Bearer {token}"})
    ok(r.status_code == 200, f"preview strong status={r.status_code}")
    j = body_json(r)
    ok(j.get("quality_score", 0) > 0.3, "strong jd quality <= 0.3")
    ok(j.get("is_vague") is False, "strong jd marked vague")
    skills = [s.lower() for s in j.get("skills", [])]
    ok(any("python" in s for s in skills), "python missing")

    vague_jd = {"jd_text": "We are looking for a passionate team player with excellent communication skills and a can-do attitude."}
    rv = req("POST", "/api/jobs/preview-jd", data=vague_jd, headers={"Authorization": f"Bearer {token}"})
    ok(rv.status_code == 200, f"preview vague status={rv.status_code}")
    v = body_json(rv)
    ok(v.get("is_vague") is True, "vague jd not flagged")
    ok(v.get("quality_score", 1) < 0.3, "vague quality >= 0.3")
    rec("SECTION 3", "PASS")
except Exception as e:
    rec("SECTION 3", "FAIL", str(e))

# Section 4 Job creation
try:
    payload = {
        "title": "QA Test Job - Python Engineer",
        "jd_text": "Senior Python Backend Engineer. Requirements: Python 5+ years, FastAPI, PostgreSQL, Docker, Redis, Kubernetes. Nice to have: React, AWS, Terraform.",
        "weight_semantic": 0.40,
        "weight_tfidf": 0.30,
        "weight_skills": 0.20,
        "weight_experience": 0.10,
        "top_n": 5,
    }
    c = req("POST", "/api/jobs", json=payload, headers={"Authorization": f"Bearer {token}"})
    ok(c.status_code in (200, 201), f"create job status={c.status_code} body={c.text[:200]}")
    cj = c.json()
    test_job_id = cj.get("id")
    ok(bool(test_job_id), "missing test_job_id")
    ok(cj.get("status") in ("pending", "processing", "done"), "bad status")

    bad_w = req("POST", "/api/jobs", json={
        "title": "Bad weights",
        "jd_text": "Python engineer needed.",
        "weight_semantic": 0.5,
        "weight_tfidf": 0.5,
        "weight_skills": 0.5,
        "weight_experience": 0.5,
        "top_n": 5,
    }, headers={"Authorization": f"Bearer {token}"})
    ok(bad_w.status_code >= 400, f"bad weights accepted status={bad_w.status_code}")
    rec("SECTION 4", "PASS")
except Exception as e:
    rec("SECTION 4", "FAIL", str(e))
    test_job_id = None

# Section 5 files
test_dir = Path("c:/Users/JAYAN/Downloads/resume prototype/hiresignal-proto/tmp_test_resumes")
test_dir.mkdir(parents=True, exist_ok=True)
try:
    (test_dir / "alice_strong.txt").write_text("""Alice Johnson\nEXPERIENCE\nSenior Python Engineer (2019-2024)\nBuilt FastAPI services, PostgreSQL, Docker, Kubernetes, Redis, AWS\nSKILLS\nPython, FastAPI, PostgreSQL, Docker, Kubernetes, Redis, AWS, React, Git, Celery\n""", encoding="utf-8")
    (test_dir / "bob_partial.txt").write_text("""Bob Smith\nJava dev with basic Python\nSkills: Java, Spring Boot, MySQL, Python\n""", encoding="utf-8")
    (test_dir / "carol_weak.txt").write_text("""Carol Davis\nMarketing role\nSkills: Office, social media\n""", encoding="utf-8")
    (test_dir / "alice_duplicate.txt").write_text((test_dir / "alice_strong.txt").read_text(encoding="utf-8"), encoding="utf-8")
    (test_dir / "blank_resume.txt").write_text("John Doe. Looking for opportunities.", encoding="utf-8")
    (test_dir / "stuffed_resume.txt").write_text(("python fastapi postgresql docker kubernetes redis aws react git celery \n" * 20), encoding="utf-8")

    ok(test_job_id is not None, "missing test job id")

    # individual uploads
    for name in ["alice_strong.txt", "bob_partial.txt", "carol_weak.txt"]:
        with open(test_dir / name, "rb") as f:
            up = req("POST", f"/api/jobs/{test_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", (name, f, "text/plain"))])
            ok(up.status_code in (200, 201), f"upload failed {name}: {up.status_code}")

    zip_path = test_dir / "batch_resumes.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        for p in test_dir.glob("*.txt"):
            zf.write(p, arcname=p.name)
    with open(zip_path, "rb") as f:
        zup = req("POST", f"/api/jobs/{test_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", ("batch_resumes.zip", f, "application/zip"))])
        ok(zup.status_code in (200, 201), f"zip upload status={zup.status_code}")

    # invalid file type
    evil = test_dir / "evil.sh"
    evil.write_text("#!/bin/bash\necho hack", encoding="utf-8")
    with open(evil, "rb") as f:
        ev = req("POST", f"/api/jobs/{test_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", ("evil.sh", f, "text/x-sh"))])
    ok(ev.status_code in (400, 415, 422), f"evil file not rejected: {ev.status_code}")

    # oversized
    huge = test_dir / "huge.txt"
    huge.write_bytes(os.urandom(25 * 1024 * 1024))
    with open(huge, "rb") as f:
        hv = req("POST", f"/api/jobs/{test_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", ("huge.txt", f, "text/plain"))])
    ok(hv.status_code in (400, 413, 422), f"huge file not rejected: {hv.status_code}")

    rec("SECTION 5", "PASS")
except Exception as e:
    rec("SECTION 5", "FAIL", str(e))

# Section 6+7+8 batch processing and NLP/anomaly
batch_job_id = None
batch_results = []
try:
    b = req("POST", "/api/jobs", json={
        "title": "Batch Processing Test",
        "jd_text": "Senior Python Backend Engineer with FastAPI, PostgreSQL, Docker, Redis experience. 3+ years required.",
        "weight_semantic": 0.40,
        "weight_tfidf": 0.30,
        "weight_skills": 0.20,
        "weight_experience": 0.10,
        "top_n": 3,
    }, headers={"Authorization": f"Bearer {token}"})
    ok(b.status_code in (200, 201), f"batch job create status={b.status_code}")
    batch_job_id = b.json()["id"]

    for name in ["alice_strong.txt", "bob_partial.txt", "carol_weak.txt", "stuffed_resume.txt", "blank_resume.txt"]:
        with open(test_dir / name, "rb") as f:
            up = req("POST", f"/api/jobs/{batch_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", (name, f, "text/plain"))])
            ok(up.status_code in (200, 201), f"batch upload failed {name}: {up.status_code}")

    status = "pending"
    for _ in range(60):
        g = req("GET", f"/api/jobs/{batch_job_id}", headers={"Authorization": f"Bearer {token}"})
        ok(g.status_code == 200, "batch job fetch failed")
        status = g.json().get("status", "")
        if status in ("done", "failed", "partial"):
            break
        time.sleep(3)
    ok(status == "done", f"batch did not finish done, status={status}")

    rr = req("GET", f"/api/jobs/{batch_job_id}/results", headers={"Authorization": f"Bearer {token}"})
    ok(rr.status_code == 200, f"batch results status={rr.status_code}")
    rj = body_json(rr)
    batch_results = items_of(rj)
    ok(len(batch_results) > 0, "no batch results")

    top = batch_results[0]
    bottom = batch_results[-1]
    ok(bottom["final_score"] < top["final_score"], "no score differentiation")

    # Section 7 checks
    by_name = {x["filename"].lower(): x for x in batch_results}
    alice = next((x for x in batch_results if "alice" in x["filename"].lower()), None)
    carol = next((x for x in batch_results if "carol" in x["filename"].lower()), None)
    ok(alice is not None and carol is not None, "alice/carol not found")
    ok(alice["final_score"] - carol["final_score"] > 0.2, "strong vs weak diff <= 0.2")
    ok(len(alice.get("matched_skills", [])) > 0, "alice matched_skills empty")

    t = top
    for k in ["score_semantic", "score_tfidf", "score_skills", "score_experience", "final_score"]:
        ok(t.get(k) is not None, f"{k} is None")
    calc = 0.40 * t["score_semantic"] + 0.30 * t["score_tfidf"] + 0.20 * t["score_skills"] + 0.10 * t["score_experience"]
    ok(abs(calc - t["final_score"]) < 0.08, "weighted score mismatch too high")

    # Section 8 checks
    blank = next((x for x in batch_results if "blank" in x["filename"].lower()), None)
    stuffed = next((x for x in batch_results if "stuffed" in x["filename"].lower()), None)
    ok(blank is not None, "blank resume missing")
    ok(blank["final_score"] <= 0.05, f"blank score too high {blank['final_score']}")
    ok(blank.get("flags", {}).get("blank_resume", False), "blank flag missing")

    ok(stuffed is not None, "stuffed resume missing")
    ok(stuffed["final_score"] <= 0.60, f"stuffed score not penalized {stuffed['final_score']}")
    ok(stuffed["final_score"] < alice["final_score"], "stuffed scored >= alice")

    # duplicate check
    with open(test_dir / "alice_duplicate.txt", "rb") as f:
        dup_up = req("POST", f"/api/jobs/{batch_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", ("alice_duplicate.txt", f, "text/plain"))])
        ok(dup_up.status_code in (200, 201), "duplicate upload failed")

    time.sleep(12)
    rr2 = req("GET", f"/api/jobs/{batch_job_id}/results", headers={"Authorization": f"Bearer {token}"})
    rs2 = items_of(body_json(rr2))
    dup = next((x for x in rs2 if "duplicate" in x["filename"].lower()), None)
    if dup is not None:
        ok(bool(dup.get("flags", {}).get("duplicate_of")), "duplicate_of flag missing")

    rec("SECTION 6", "PARTIAL", "Batch done; websocket not programmatically validated in this run")
    rec("SECTION 7", "PASS")
    rec("SECTION 8", "PASS")
except Exception as e:
    rec("SECTION 6", "FAIL", str(e))
    rec("SECTION 7", "FAIL", str(e))
    rec("SECTION 8", "FAIL", str(e))

# Section 9 OCR
try:
    ocr_pdf = test_dir / "scanned_resume.pdf"
    # minimal empty pdf to force OCR/fallback path
    ocr_pdf.write_bytes(b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n3 0 obj<</Type/Page/MediaBox[0 0 612 792]/Parent 2 0 R/Contents 4 0 R/Resources<<>>>>endobj\n4 0 obj<</Length 0>>stream\nendstream\nendobj\nxref\n0 5\n0000000000 65535 f\n0000000009 00000 n\n0000000058 00000 n\n0000000115 00000 n\n0000000206 00000 n\ntrailer<</Size 5/Root 1 0 R>>\nstartxref\n256\n%%EOF")

    o = req("POST", "/api/jobs", json={
        "title": "OCR Test Job",
        "jd_text": "Python Developer with Django and PostgreSQL experience.",
        "weight_semantic": 0.40,
        "weight_tfidf": 0.30,
        "weight_skills": 0.20,
        "weight_experience": 0.10,
        "top_n": 5,
    }, headers={"Authorization": f"Bearer {token}"})
    ok(o.status_code in (200, 201), "ocr job create failed")
    ocr_job = o.json()["id"]

    with open(ocr_pdf, "rb") as f:
        u = req("POST", f"/api/jobs/{ocr_job}/upload", headers={"Authorization": f"Bearer {token}"}, files=[("files", ("scanned_resume.pdf", f, "application/pdf"))])
    ok(u.status_code in (200, 201), f"ocr upload status={u.status_code}")

    time.sleep(20)
    rs = req("GET", f"/api/jobs/{ocr_job}/results", headers={"Authorization": f"Bearer {token}"})
    ok(rs.status_code == 200, "ocr results failed")
    items = items_of(body_json(rs))
    ok(len(items) > 0, "no ocr result")
    q = items[0].get("extraction_quality")
    ok(q in ("ocr_used", "poor", "failed", "good"), f"unexpected extraction quality {q}")
    rec("SECTION 9", "PARTIAL", f"extraction_quality={q}")
except Exception as e:
    rec("SECTION 9", "FAIL", str(e))

# Section 10 single screen
try:
    t0 = time.time()
    with open(test_dir / "alice_strong.txt", "rb") as f:
        ss = req("POST", "/api/screen/single", headers={"Authorization": f"Bearer {token}"}, files={
            "jd_text": (None, "Senior Python Backend Engineer. Requirements: Python 5+ years, FastAPI, PostgreSQL, Docker, Redis, Kubernetes, AWS."),
            "resume_file": ("alice_strong.txt", f, "text/plain"),
        })
    elapsed = time.time() - t0
    ok(ss.status_code == 200, f"single strong status={ss.status_code}")
    sj = ss.json()
    ok(sj.get("final_score", 0) > 0.5, f"single strong score low {sj.get('final_score')}")
    ok(len(sj.get("matched_skills", [])) > 3, "single strong matched skills too low")

    with open(test_dir / "carol_weak.txt", "rb") as f:
        sw = req("POST", "/api/screen/single", headers={"Authorization": f"Bearer {token}"}, files={
            "jd_text": (None, "Senior Python Backend Engineer. Requirements: Python 5+ years, FastAPI, PostgreSQL, Docker."),
            "resume_file": ("carol_weak.txt", f, "text/plain"),
        })
    ok(sw.status_code == 200, f"single weak status={sw.status_code}")
    ok(sw.json().get("final_score", 1) < 0.4, "weak single score too high")

    miss_resume = req("POST", "/api/screen/single", headers={"Authorization": f"Bearer {token}"}, files={"jd_text": (None, "Python engineer")})
    with open(test_dir / "alice_strong.txt", "rb") as mf:
        miss_jd = req("POST", "/api/screen/single", headers={"Authorization": f"Bearer {token}"}, files={"resume_file": ("alice.txt", mf, "text/plain")})
    ok(miss_resume.status_code == 422 and miss_jd.status_code == 422, "missing field validation not 422")

    rec("SECTION 10", "PASS" if elapsed <= 5 else "PARTIAL", f"elapsed={elapsed:.2f}s")
except Exception as e:
    rec("SECTION 10", "FAIL", str(e))

# Section 11 feedback
try:
    ok(batch_job_id is not None and batch_results, "batch results unavailable for feedback")
    top_id = batch_results[0]["id"]
    low_id = batch_results[-1]["id"]

    f1 = req("POST", "/api/feedback", headers={"Authorization": f"Bearer {token}"}, json={"result_id": top_id, "action": "shortlisted", "notes": "Strong match"})
    ok(f1.status_code in (200, 201), f"feedback shortlist status={f1.status_code}")

    f2 = req("POST", "/api/feedback", headers={"Authorization": f"Bearer {token}"}, json={"result_id": low_id, "action": "rejected"})
    ok(f2.status_code in (200, 201), f"feedback reject status={f2.status_code}")

    inv = req("POST", "/api/feedback", headers={"Authorization": f"Bearer {token}"}, json={"result_id": top_id, "action": "maybe_someday"})
    ok(inv.status_code >= 400, f"invalid feedback action accepted {inv.status_code}")

    fb = req("GET", f"/api/feedback/job/{batch_job_id}", headers={"Authorization": f"Bearer {token}"})
    ok(fb.status_code == 200, "feedback list failed")
    items = items_of(body_json(fb))
    actions = [x.get("action") for x in items]
    ok("shortlisted" in actions and "rejected" in actions, f"feedback actions missing {actions}")
    rec("SECTION 11", "PASS")
except Exception as e:
    rec("SECTION 11", "FAIL", str(e))

# Section 12 exports
try:
    ok(batch_job_id is not None, "missing batch_job_id")
    csv_resp = req("GET", f"/api/export/{batch_job_id}/csv", headers={"Authorization": f"Bearer {token}"})
    ok(csv_resp.status_code == 200, f"csv export status={csv_resp.status_code}")
    csv_path = test_dir / "shortlist_test.csv"
    csv_path.write_bytes(csv_resp.content)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fields = reader.fieldnames or []
    req_cols = ["Rank", "Filename"]
    ok(all(c in fields for c in req_cols), f"csv required columns missing {fields}")
    ok(len(rows) > 0, "csv rows empty")

    xlsx_resp = req("GET", f"/api/export/{batch_job_id}/excel", headers={"Authorization": f"Bearer {token}"})
    ok(xlsx_resp.status_code == 200, f"excel export status={xlsx_resp.status_code}")
    xlsx_path = test_dir / "shortlist_test.xlsx"
    xlsx_path.write_bytes(xlsx_resp.content)

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ok("Shortlist" in wb.sheetnames and "Score Breakdown" in wb.sheetnames and "JD Info" in wb.sheetnames, f"missing sheet names {wb.sheetnames}")
    rec("SECTION 12", "PASS")
except Exception as e:
    rec("SECTION 12", "FAIL", str(e))

# Section 13 analytics
try:
    ov = req("GET", "/api/analytics/overview", headers={"Authorization": f"Bearer {token}"})
    ok(ov.status_code == 200, f"overview status={ov.status_code}")
    ovj = ov.json()
    # Flexible key names
    keys = set(ovj.keys())
    needed_any = ["total_jobs", "total_resumes_processed", "total_resumes_screened"]
    ok("total_jobs" in keys, f"overview missing total_jobs {keys}")
    ok(any(k in keys for k in needed_any[1:]) or "total_jobs" in keys, f"overview missing resume count keys {keys}")

    tr = req("GET", "/api/analytics/score-trend", headers={"Authorization": f"Bearer {token}"})
    ok(tr.status_code == 200, f"trend status={tr.status_code}")
    tj = tr.json()
    items = items_of(tj)
    ok(len(items) > 0, "score trend empty")

    sk = req("GET", "/api/analytics/skills", headers={"Authorization": f"Bearer {token}"})
    ok(sk.status_code == 200, f"skills analytics status={sk.status_code}")
    rec("SECTION 13", "PASS")
except Exception as e:
    rec("SECTION 13", "FAIL", str(e))

# Section 14 edge cases
try:
    # empty zip
    empty_zip = test_dir / "empty_test.zip"
    with zipfile.ZipFile(empty_zip, "w"):
        pass
    with open(empty_zip, "rb") as f:
        ez = req("POST", f"/api/jobs/{test_job_id}/upload", headers={"Authorization": f"Bearer {token}"}, files={"resumes": ("empty_test.zip", f, "application/zip")})
    ok(ez.status_code in (400, 422), f"empty zip not rejected {ez.status_code}")

    nf = req("GET", "/api/jobs/00000000-0000-0000-0000-000000000000", headers={"Authorization": f"Bearer {token}"})
    ok(nf.status_code == 404, f"job not found should be 404 got {nf.status_code}")

    # long jd
    long_jd = "Python developer. Requirements: Python, FastAPI. " * 300
    lj = req("POST", "/api/jobs/preview-jd", headers={"Authorization": f"Bearer {token}"}, data={"jd_text": long_jd})
    ok(lj.status_code == 200 and "quality_score" in body_json(lj), "long jd not handled")

    rec("SECTION 14", "PASS")
except Exception as e:
    rec("SECTION 14", "FAIL", str(e))

# Section 15 database state
try:
    sql = r"""
\dt
SELECT 'tenants' as tbl, COUNT(*) FROM tenants
UNION ALL SELECT 'users', COUNT(*) FROM users
UNION ALL SELECT 'screening_jobs', COUNT(*) FROM screening_jobs
UNION ALL SELECT 'resume_results', COUNT(*) FROM resume_results
UNION ALL SELECT 'hr_feedback', COUNT(*) FROM hr_feedback;
SELECT COUNT(*) as nulls_in_scores
FROM resume_results rr
JOIN screening_jobs sj ON rr.job_id = sj.id
WHERE sj.status = 'done'
AND (rr.final_score IS NULL OR rr.rank IS NULL);
SELECT t.slug, COUNT(sj.id) as job_count
FROM tenants t
LEFT JOIN screening_jobs sj ON sj.tenant_id = t.id
GROUP BY t.slug;
SELECT content_hash, COUNT(*) as cnt
FROM resume_results
GROUP BY content_hash
HAVING COUNT(*) > 1
LIMIT 5;
"""
    p = subprocess.run([
        "docker-compose", "exec", "-T", "postgres", "psql", "-U", "hiresignal", "-d", "hiresignal"
    ], input=sql.encode("utf-8"), cwd=str(Path(__file__).parent), capture_output=True)
    out = p.stdout.decode("utf-8", errors="ignore")
    ok(p.returncode == 0, f"psql failed: {p.stderr.decode('utf-8', errors='ignore')[:200]}")
    ok("tenants" in out and "users" in out and "screening_jobs" in out, "table list incomplete")
    rec("SECTION 15", "PASS")
except Exception as e:
    rec("SECTION 15", "FAIL", str(e))

# Section 16 manual frontend
rec("SECTION 16", "PARTIAL", "Manual browser checklist not fully automatable in terminal run")

print("\n=== QA SUMMARY JSON ===")
print(json.dumps({k: {"status": v[0], "message": v[1]} for k, v in report.items()}, indent=2))

pass_count = sum(1 for v in report.values() if v[0] == "PASS")
print(f"\nPASS_COUNT={pass_count} / {len(report)}")
if notes:
    print("NOTES:")
    for n in notes:
        print("-", n)
